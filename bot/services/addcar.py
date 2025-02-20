import re
import threading

import telebot
import openpyxl
from telebot import types
from django.db.models import Q
from django.utils import timezone
from telebot.types import InlineKeyboardButton, InlineKeyboardMarkup
from openpyxl.utils import get_column_letter

from conf.settings import ADMINS, CHANNEL_ID

from ..buttons.default import ask_phone, main_button, main_menu, main_button2
from ..buttons.inline import create_social_btn, urlkb
from ..models import Car, CarImage, Search, TgUser, Region, District
from ..services.steps import USER_STEP
from ..utils.post_car_photo import post_photo_to_telegraph
from ..utils.search import search_cars

phone_number_pattern = r'^(\+998|0)[1-9]\d{8}$'


def is_phone_number(value):
    if re.match(phone_number_pattern, value):
        return True
    else:
        return False


def add_car(message, bot):
    try:
        user = TgUser.objects.get(telegram_id=message.from_user.id)
        if message.photo:

            car, created = Car.objects.get_or_create(
                owner=user, complate=False)
            

            if car.images.count() < 6:
                photo_info = bot.get_file(message.photo[-1].file_id)
                photo = bot.download_file(photo_info.file_path)

                car_image = CarImage.objects.create(
                    car=car, image_link=message.photo[-1].file_id)

                # t1 = threading.Thread(target=post_photo_to_telegraph,
                #                       args=(car_image, photo))
                # t1.start()

            if car.delete or not created:
                bot.delete_message(
                    chat_id=message.from_user.id, message_id=car.delete)
            
            if not car.name:
                msg = bot.send_message(
                    message.from_user.id, text='🚘 Lyustra modelini kiriting\n(<i>Karobkada yozilgan, Kamida 5 ta belgi</i>)\nMisol: "NC62/2W/2 CH"', parse_mode='html')
            elif car.name and not car.model:
                msg = bot.send_message(
                        message.from_user.id, text='❕Lyustra uchun kerak bo\'lgan extiyot qisimni va sonini kiriting.\n(<i>shisha, qosh Va.h.k... 2ta</i>)', parse_mode='html')
                user.step = USER_STEP['ADD_MODEL']
                user.save()
            car.delete = msg.id
            car.save()
        elif message.text and user.car_set.filter(complate=False).exists():
            car = user.car_set.get(complate=False)
            if not car.name and 50 > len(message.text) >= 5:
                car.name = message.text.capitalize()
                car.save()
                msg = bot.send_message(
                    message.from_user.id, text='📷 Lyustrada yetshmayotgan extiyot qism rasmini joylang!', parse_mode='html')
                car.delete = msg.id
                car.save()
            elif not car.name and len(message.text) < 5:
                msg = bot.send_message(
                    message.from_user.id, text='🚘 Lyustra modelini kiriting\n(<i>Karobkada yozilgan, Kamida 5 ta belgi</i>)\nMisol: "NC62/2W/2 CH"', parse_mode='html')
                if car.delete:
                    bot.delete_message(
                        chat_id=message.from_user.id, message_id=car.delete)
                car.delete = msg.id
                car.save()
            else:
                bot.delete_message(
                chat_id=message.from_user.id, message_id=message.id)
                msg = bot.send_message(
                    message.from_user.id, text='📷 Lyustrada yetshmayotgan extiyot qism rasmini joylang!', parse_mode='html')
                if car.delete:
                    bot.delete_message(
                        chat_id=message.from_user.id, message_id=car.delete)
                car.delete = msg.id
                car.save()
        else:
            msg = bot.send_message(
                message.from_user.id, text='🚫 Iltimos! <b>2</b> tadan <b>6</b> tagacha Lyustrangiz rasmini joylang!', parse_mode='html')
    except Exception as e:
        print(e)
        pass


def add_model(message, bot):
    if message.text:
        model = message.text.capitalize()
        if 50 > len(model) >= 5:
            user = TgUser.objects.get(telegram_id=message.from_user.id)
            car = Car.objects.get(owner=user, complate=False)
            car.model = model
            car.save()
            user.step = USER_STEP['ADD_YEAR']
            user.save()
            bot.send_message(message.from_user.id,
                            text=f'📅 Lyustrani xarid qilgan sanangizni kiriting.\n<i>(25.01.2025)</i>', parse_mode='html')
        else:
            bot.send_message(message.from_user.id,
                            text='🚫 Lyustra zapchasti xaqida batafsilroq ma\'lumot kiriting!\n(<i>shisha, qosh Va.h.k... 2ta</i>)', parse_mode='html')
    else:
        bot.send_message(message.from_user.id,
                            text='🚫 Lyustra zapchasti xaqida batafsilroq ma\'lumot kiriting!\n(<i>shisha, qosh Va.h.k... 2ta</i>)', parse_mode='html')

def add_year(message, bot):
    try:
        if len(message.text) >= 2:
            year = message.text
            user = TgUser.objects.get(telegram_id=message.from_user.id)
            car = Car.objects.get(owner=user, complate=False)
            car.year = year
            car.save()
            user.step = USER_STEP['ADD_PRICE']
            user.save()
            regions = Region.objects.all()
            keyboard = types.InlineKeyboardMarkup()
            counter = 0
            for region in regions:
                if counter % 2 == 0:
                    # Create a new row for every two regions
                    row = []

                button = types.InlineKeyboardButton(
                    text=region.name, callback_data='region_' + str(region.id))
                row.append(button)

                if counter % 2 == 1 or counter == len(regions) - 1:
                    # Add the row to the keyboard if two regions are added or it's the last region
                    keyboard.row(*row)

                counter += 1
            bot.send_message(message.from_user.id,
                             text='Lyustra Xarid qilingan viloyatni tanlang! 👇', parse_mode='html', reply_markup=keyboard)
        else:
            bot.send_message(
                message.from_user.id, text=f'🚫Lyustra xarid qilingan yil qabul qilinmadi.\nQayta kiriting! <i>(1999-{timezone.now().year}</i>)', parse_mode='html')
    except Exception as e:
        print(e)
        bot.send_message(
            message.from_user.id, text=f'🚫Lyustra xarid qilingan yil qabul qilinmadi.\nQayta kiriting! <i>(1999-{timezone.now().year})</i>', parse_mode='html')


def add_price(message, bot):
    try: 
        bot.delete_message(chat_id=message.chat.id, message_id=message.id)
        bot.send_message(
            message.from_user.id, text='Viloyat va tumanni tanlang 👆', parse_mode='html')
    except Exception as e:
        print(e)
        bot.send_message(
            message.from_user.id, text='Viloyat va tumanni tanlang 👆', parse_mode='html')


def add_description(message, bot):
    try:
        description = message.text
        if 10 < len(description) < 800:
            user = TgUser.objects.get(telegram_id=message.from_user.id)
            car = Car.objects.get(owner=user, complate=False)
            car.description = description
            car.save()
            user.step = USER_STEP['ADD_NUMBER']
            user.save()
            bot.send_message(message.from_user.id, text='Bog\'lanish uchun telefon raqamingizni kiriting',
                             reply_markup=ask_phone, parse_mode='html')
        else:
            bot.send_message(
                message.from_user.id, text='🚫Lyustra haqida ma\'lumot qabul qilinmadi.\nQayta kiriting', parse_mode='html')
    except:
        bot.send_message(
            message.from_user.id, text='🚫Lyustra haqida ma\'lumot qabul qilinmadi.\nQayta kiriting', parse_mode='html')


def add_number(message, bot):
    try:
        if message.content_type == 'text':
            if is_phone_number(message.text):
                contact_number = message.text
                user = TgUser.objects.get(telegram_id=message.from_user.id)
                car = Car.objects.get(owner=user, complate=False)
                car.contact_number = contact_number
                car.complate = True
                car.save()
                user.step = USER_STEP['DEFAULT']
                if user.phone == '-':
                    user.phone = contact_number
                user.save()
                bot.send_message(message.from_user.id, text='Ariza muvofaqiyatli joylandi',
                                 reply_markup=main_button2, parse_mode='html')

            else:
                bot.send_message(
                    message.from_user.id, text='🚫 Iltimos telefon raqamini to\'g\'ri farmatda kiriting.', parse_mode='html')
        elif message.content_type == 'contact':
            contact_number = message.contact.phone_number
            user = TgUser.objects.get(telegram_id=message.from_user.id)
            car = Car.objects.get(owner=user, complate=False)
            car.contact_number = contact_number
            car.complate = True
            car.save()
            user.step = USER_STEP['DEFAULT']
            user.phone = contact_number
            user.save()
            bot.send_message(message.from_user.id, text='Ariza muvofaqiyatli joylandi',
                                 reply_markup=main_button2, parse_mode='html')

        if car:
            # send new car to admins and channel
            text = f"Nomi: {car.name},\nExtiyot qisim: {car.model},\nIshlab chiqarilgan yil: {car.year},\nViloyat: {car.region.name},\nTuman: {car.district.name},\nQo'shimcha malumot: \n{car.description[:800]},\n\nBog'lanish: {car.contact_number}\n\n"
            text += f"Joylandi: {car.created_at.strftime('%Y-%m-%d')}"
            media_group = [telebot.types.InputMediaPhoto(
                media=car.images.first().image_link, caption=text)]
            for photo in car.images.all()[1:]:
                media_group.append(
                    telebot.types.InputMediaPhoto(media=photo.image_link))
            for admin in ADMINS:
                try:
                    msg = bot.send_media_group(
                        chat_id=admin, media=media_group)
                    ids = ''
                    for a in msg:
                        ids += ','+str(a.id)
                    bot.reply_to(message=msg[0], text="Ushbu arizani o\'chirish", reply_markup=InlineKeyboardMarkup(row_width=1).add(
                        InlineKeyboardButton(text='Faollashtirish', callback_data=f'post_{car.id}'+ids),))
                except:
                    pass

            # send to channel
            # bot.send_media_group(
            #     chat_id=CHANNEL_ID, media=media_group)

            # send to the owner
            if message.from_user.id not in ADMINS:

                msg = bot.send_media_group(
                    chat_id=message.from_user.id, media=media_group)

                ids = ''
                for a in msg:
                    ids += ','+str(a.id)

                # bot.reply_to(message=msg[0], text="Ushbu arizani o\'chirish", reply_markup=InlineKeyboardMarkup().add(
                #     InlineKeyboardButton(text=f'O\'chirish', callback_data=f'del_{car.id}'+ids)))

                msg = bot.send_message(message.from_user.id, text='✅ Ariza qabul qilindi!\nArizangiz tez orada ko\'rib chiqilib Adminlar tomonidan faollashtiriladi!',
                                       reply_markup=main_button, parse_mode='html')

                bot.reply_to(message=msg, text="Murojat uchun👇",
                             reply_markup=urlkb)

    except Exception as e:
        print(e)
        # bot.send_message(
        #     message.from_user.id, text='Iltimos telefon raqamini to\'g\'ri farmatda kiriting.', parse_mode='html')


def get_serach_result(text, user_id):
    search_text = text
    user = TgUser.objects.get(telegram_id=user_id)

    cars = []
    if search_text == '/all':
        cars = Car.objects.filter(complate=True).order_by('-created_at')
    elif 50 > len(search_text) >= 3:
        cars = search_cars(search_text)
    if len(cars) > 2:
        search = Search.objects.create(text=search_text, user=user)
        search_id = search.id
    else:
        search_id = 0

    return {'cars': cars, 'search_id': search_id}


def paginated(text):
    search_text = text
    cars = []
    if search_text == '/all':
        cars = Car.objects.filter(complate=True).order_by('-created_at')
    elif 50 > len(search_text) > 3:
        cars = search_cars(search_text)
    return cars


def search_car(message, bot):
    result = get_serach_result(
        text=message.text, user_id=message.from_user.id)
    cars = result['cars']
    search_id = result['search_id']
    search_text = message.text
    if 2 >= len(cars) > 0:
        TgUser.objects.filter(telegram_id=message.from_user.id).update(
            step=USER_STEP['DEFAULT'])
        bot.send_message(chat_id=message.from_user.id,
                         text='Natijalar', reply_markup=main_button)
        for car in cars:
            car.seen.add(TgUser.objects.get(telegram_id=message.from_user.id))
            car_id = car.id
            text = f"Nomi: {car.name},\nExtiyot qisim: {car.model},\nIshlab chiqarilgan yil: {car.year},\nViloyat: {car.region.name},\nTuman: {car.district.name},\nQo'shimcha malumot: \n{car.description[:800]},\n\nBog'lanish: {car.contact_number}\n\n"
            text += f"👁: {car.seen.count()}, "
            text += f"👍: {car.likes.count()}, "
            text += f"👎: {car.dislikes.count()}\n\n"
            text += f"Joylandi: {car.created_at.strftime('%Y-%m-%d')}"
            media_group = [telebot.types.InputMediaPhoto(
                media=car.images.first().image_link, caption=text)]
            for photo in car.images.all()[1:]:
                media_group.append(
                    telebot.types.InputMediaPhoto(media=photo.image_link))

            msg = bot.send_media_group(
                chat_id=message.from_user.id, media=media_group)

            bot.reply_to(
                message=msg[0], text="Ariza xaqida fikir bildirasizmi?", reply_markup=create_social_btn(car_id))
    elif len(cars) > 2:
        TgUser.objects.filter(telegram_id=message.from_user.id).update(
            step=USER_STEP['DEFAULT'])
        bot.send_message(chat_id=message.from_user.id,
                         text='Natijalar', reply_markup=main_button)
        inline_kb = InlineKeyboardMarkup(row_width=5)
        buttons = []
        text = f"<strong>{search_text}</strong> so'rovi bo'yicha natijalar:\n{len(cars)} dan 1 - {10 if len(cars) >= 10 else len(cars)}\n\n"
        text += "<pre>"
        text += "{:<2} {:<11} {:<6} {:<9}\n\n".format(
            "No", "Nomi", "Yili", "Viloyat")
        for count, car in enumerate(cars[:10]):
            text += "{:<2} {:<11} {:<6} {:<9}\n".format(
                str(count+1)+".", car.name[:8]+'...' if len(car.name) > 11 else car.name, car.year, car.region.name)
            button = InlineKeyboardButton(
                text=str(count+1), callback_data=f"retrieve_{car.id}")
            buttons.append(button)

        text += "</pre>"
        inline_kb.add(*buttons)
        inline_kb.add(InlineKeyboardButton(f'⬅', callback_data=f'prev {search_id}'),
                      InlineKeyboardButton(
                          f'❌', callback_data=f'remove {search_id}'),
                      InlineKeyboardButton(f'➡', callback_data=f'next {search_id}'))
        bot.send_message(message.from_user.id, text,
                         parse_mode='html', reply_markup=inline_kb)

    else:
        bot.send_message(message.from_user.id, text="So'rov bo'yicha xechqanday Ariza topilmadi",
                         parse_mode='html', reply_markup=main_menu)


def create_excel_report(reports, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reports"

    # Define the header
    headers = ['ID', 'Lyustra Nomi', 'Yetishmayotgan extiyot qisim', 'Viloyat', 'Tuman', 'Telefon raqami', 'Holati', 'Joylangan sana']
    ws.append(headers)

    # Populate the sheet with data from the reports
    for report in reports:
        row = [
            report.id,
            report.name,
            report.model, 
            report.region.name,
            report.district.name,
            report.contact_number,
            report.post,
            report.created_at.strftime('%d.%m.%Y')
        ]
        ws.append(row)
        
        # Auto-size columns based on the length of the data in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column number
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # Add a little extra space
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Save the workbook to the specified file path
    wb.save(file_path)
    wb.close()